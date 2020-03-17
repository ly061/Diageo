"""
The market matrix excel utils
Get some matrix data from excel file. "MY20-RYI-Matrix-V3.1.xlsx"
"""

__author__ = "Michael.Tian"

import openpyxl
from bs4 import BeautifulSoup
from requests.adapters import HTTPAdapter
from selenium import webdriver
import requests

# __matrix_file_name__ = "proj21_escape_everyday/data/V2.xlsx"
__matrix_file_name__ = "V2.xlsx"

def get_social_matrix():
    """
    Get all soical link matrix
    locale column B
    Facebook Url column W
    Instagram Url column X
    Twitter Url column Y
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_file_name__)
    sheet = wb["Sheet1"]
    res = {}
    for index in range(5, 46):
        locale = sheet["C{}".format(index)].value
        facebook = sheet["BA{}".format(index)].value
        instagram = sheet["AZ{}".format(index)].value
        twitter = sheet["AY{}".format(index)].value
        Youtube = sheet["AX{}".format(index)].value
        res[locale] = [link for link in [facebook, instagram, twitter, Youtube] if link]

    wb.close()
    return res

def get_all_locale():
    """
    Get all locale from matrix
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_file_name__)
    sheet = wb["RYI"]

    res = []
    for index in range(4, 35):
        locale = sheet["B{}".format(index)].value.split()[0]
        res.append(locale)

    return res

def get_bike_matrix():
    """
    Get All bike info matrix
    Custom: K,L,M,N,O
    Performance: P,Q,R,S
    Touring: T,U,V
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_file_name__)
    sheet = wb["Sheet1"]
    bike_category = {
        0: ["K", "L", "M", "N", "O"],
        1: ["P", "Q", "R", "S"],
        2: ["T", "U", "V"]
    }

    bike_column = ["K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V", "w", "x", "Y", "Z", "AA", "AB", "AC", "AD", "AE", "AF", "AG", "AH", "AI", "AJ", "AK", "AL", "AM", "AN"]
    # bike_column = [col for c in [bike_category[key] for key in bike_category.keys()] for col in c]
    bike_code = {}
    for col in bike_column:
        code = sheet["{}4".format(col)].value.replace("\n", "")  # 获取到每一列的车名
        bike_code[col] = code

    res = {}
    for index in range(5, 46):
        locale = sheet["C{}".format(index)].value.split()[0]  # 获取 local
        res[locale] = []
        for key in bike_column:
            # res[locale][key] = []
            # column = bike_category[key]
            # for col in column:
            #     cc = sheet["{}{}".format(col, index)].value
            #     if cc and not cc.startswith("N"):
            #         res[locale][key].append(bike_code[col])

            cc = sheet["{}{}".format(key, index)].value
            if cc and cc.startswith(("Y", "y")):
                res[locale].append(bike_code[key])          # 获取到每个locale有哪些车

    wb.close()
    return res



if __name__ == "__main__":
    get_bike_matrix()